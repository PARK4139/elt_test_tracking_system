from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.services.test_result_service import list_recent_test_results


def _datetime_to_isoformat(value):
    if value is None:
        return None
    return value.strftime("%Y-%m-%d %H:%M:%S")


def build_test_result_workbook(database_session: Session) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "test_results"

    header_names = [
        "양식제출ID",
        "데이터생성시각",
        "업체명",
        "양식제출자",
        "모델명",
        "공정번호",
        "월",
        "검사대수",
        "저온 투입일",
        "저온 완료일",
        "저온 시간",
        "고온 투입일",
        "고온 완료일",
        "고온 시간",
        "데이터수정시각",
    ]
    worksheet.append(header_names)

    for test_result in list_recent_test_results(database_session=database_session, limit=1000):
        worksheet.append(
            [
                test_result.form_submission_id,
                _datetime_to_isoformat(test_result.created_at),
                test_result.key_1,
                test_result.key_2,
                test_result.key_3,
                test_result.key_4,
                test_result.field_01,
                test_result.field_02,
                _datetime_to_isoformat(test_result.low_test_started_at),
                _datetime_to_isoformat(test_result.low_test_ended_at),
                test_result.low_test_delta,
                _datetime_to_isoformat(test_result.high_test_started_at),
                _datetime_to_isoformat(test_result.high_test_ended_at),
                test_result.high_test_delta,
                _datetime_to_isoformat(test_result.updated_at),
            ]
        )

    return workbook


def append_test_results_to_existing_workbook(
    database_session: Session,
    excel_file_path: str,
    sheet_name: str,
    limit: int = 1000,
) -> dict[str, int | str]:
    """기존 엑셀 파일의 지정 시트에 현재 데이터를 append."""
    path = (excel_file_path or "").strip()
    if not path:
        raise ValueError("excel_file_path is required.")
    target_sheet = (sheet_name or "").strip()
    if not target_sheet:
        raise ValueError("sheet_name is required.")

    workbook = load_workbook(path)
    worksheet = workbook[target_sheet] if target_sheet in workbook.sheetnames else workbook.create_sheet(target_sheet)

    # If sheet is empty, write header
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        worksheet.append(
            [
                "양식제출ID",
                "데이터생성시각",
                "업체명",
                "양식제출자",
                "모델명",
                "공정번호",
                "월",
                "검사대수",
                "저온 투입일",
                "저온 완료일",
                "저온 시간",
                "고온 투입일",
                "고온 완료일",
                "고온 시간",
                "데이터수정시각",
            ]
        )

    appended = 0
    for test_result in list_recent_test_results(database_session=database_session, limit=limit):
        worksheet.append(
            [
                test_result.form_submission_id,
                _datetime_to_isoformat(test_result.created_at),
                test_result.key_1,
                test_result.key_2,
                test_result.key_3,
                test_result.key_4,
                test_result.field_01,
                test_result.field_02,
                _datetime_to_isoformat(test_result.low_test_started_at),
                _datetime_to_isoformat(test_result.low_test_ended_at),
                test_result.low_test_delta,
                _datetime_to_isoformat(test_result.high_test_started_at),
                _datetime_to_isoformat(test_result.high_test_ended_at),
                test_result.high_test_delta,
                _datetime_to_isoformat(test_result.updated_at),
            ]
        )
        appended += 1

    workbook.save(path)
    return {"sheet_name": target_sheet, "appended_rows": appended}
